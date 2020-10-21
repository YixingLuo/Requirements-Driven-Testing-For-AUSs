#https://github.com/bitcoin/bitcoin/issues?page=3&q=is%3Aissue+is%3Aclosed

import urllib.request
import re
from bs4 import BeautifulSoup
import io
import sys
import openpyxl
import xlwt
import xlrd
from xlutils.copy import copy


P=sys.argv[0]
record=[]
'''
def excelwrite(filename,value):			
	workbook = xlrd.open_workbook(filename)	
	sheet = workbook.sheet_by_index(0)	
	rowNum = sheet.nrows	
	colNum = sheet.ncols	
	newbook = copy(workbook)	
	newsheet = newbook.get_sheet(0)	# 在末尾增加新行
	for i in range(0,len(value)):
		newsheet.write(rowNum, i, value[i])	# 覆盖保存	
	newbook.save(filename)
'''
def write07Excel(path,value):
	wb = openpyxl.load_workbook(path)
	ws = wb['Sheet1']
	#for x in data:
	ws.append(value)
	wb.save(path)
	#print("写入成功")

def gettitle(page=1):
	try:
		url="https://github.com/bitcoin/bitcoin/issues?page="+str(page)+"&q=is%3Aissue+is%3Aclosed"
		data = urllib.request.urlopen(url).read()
		z_data = data.decode('UTF-8')
		soup = BeautifulSoup(z_data, 'lxml')
		a = soup.select('li > div > div > a')
		b=soup.select('span.opened-by')
		c=soup.select('relative-time')
		test=soup.select('div.float-left.col-9.lh-condensed.p-2')
		#hostsfile = open('record.txt', 'w', newline='',encoding='UTF-8')
		for i in range(0,len(b)):
			temp=[]
			temp.append(a[i].get_text())                    #标题
			temp.append("closed")                           #状态
			temp.append(c[i].attrs['datetime'])             #问题提出时间
			z=""
			for j in test[i].select('a.d-inline-block.IssueLabel.v-align-text-top'):
				z+=j.get_text()+'/'
			temp.append(z)                                  #标签
			#sn=b[i].get_text().replace(" ","").split('\n')[1].replace("#","").replace("\n","")
			m = re.search('\d+',b[i].get_text())
			n=m.group(0)
			temp.append(m.group(0))                         #任务ID
			s,t=getdata(m.group(0))
			#s=getdata(m.group(0))
			temp.append(t)                                  #任务关闭时间
			for i in s:
				temp.append(i)
			#temp.append(s)
			write07Excel("closed.xlsx",temp)
			#record.append(temp)
		#hostsfile.close()
		print('hosts刷新成功:',len(a))
	except Exception as err:
		print(str(err))

def getdata(sn):
	temp=[]
	try:
		url="https://github.com/bitcoin/bitcoin/issues/"+str(sn)
		data = urllib.request.urlopen(url).read()
		z_data = data.decode('UTF-8')
		soup = BeautifulSoup(z_data, 'lxml')
		a = soup.select('task-lists table > tbody > tr > td.d-block.comment-body.markdown-body.js-comment-body')
		b = soup.select('div.discussion-item.discussion-item-closed')
		author=soup.find_all('h3',attrs={'class':'timeline-comment-header-text f5 text-normal'})
		#print(sn,len(a),len(author))
		temp.append(len(a)-1)                      #评论数
		temp.append(author[0].select('a.author')[0].get_text())      #问题提出者
		if len(b[0].select('a.author'))>0:
			temp.append(b[0].select('a.author')[0].get_text())           #问题关闭者
		else:
			temp.append("")
		temp.append(a[0].get_text())
		for i in range(1,len(a)):
			tempd=[]
			tempd.append(sn)                                                       #问题id
			tempd.append(author[i].select('a.author')[0].get_text())               #评论人id
			tempd.append(author[i].select('relative-time')[0].attrs['datetime'])   #评论时间
			tempd.append(a[i].get_text())                                          #评论内容                                         #评论内容
			write07Excel("closed_comment.xlsx",tempd)
	except Exception as err:
		print(str(err))
		#pass
	if len(b[0].select('relative-time'))>0:
		return temp,b[0].select('relative-time')[0].attrs['datetime']
	else:
		return temp,""

'''
def write07Excel(path,value,row):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    wb.save(path)
'''

if __name__=="__main__":
	#for i in range(37,146):
	try:
		gettitle(float(P))
		print("第"+str(P)+"页完成")
	except Exception as err:
		hostsfile = open('record.txt', 'w', newline='')
		hostsfile.write(str(P)+":"+str(err) + "\n")
		hostsfile.close()
	    print("第"+str(i)+"页抓取失败")
	#write07Excel("closed.xlsx",record)